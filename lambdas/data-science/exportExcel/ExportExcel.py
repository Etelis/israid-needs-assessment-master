import os
import json
import logging
import boto3 
import glob
from botocore.exceptions import ClientError
from boto3.dynamodb.conditions import Key, Attr
from openpyxl import Workbook

## Change this global settings if necessary
REGION = 'eu-north-1'
ANSWERS_TABLE_NAME = 'Answers'
ANSWERS_RNAS_TABLE_KEY = 'rnaId'
RNAS_TABLE_NAME = 'Rnas'
RNAS_PARTITION_KEY = 'id'
CATEGORIES_JSON_PATH = "categories.json"
SUBCATEGORIES_JSON_PATH = "sub-categories.json"
QUESTIONS_JSON_PATH = "questions.json"

# Configure logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)


def get_questions() -> dict:
    # Fetching all questions from local: client\src\static-data\questions.json
    questions = json.load(open(QUESTIONS_JSON_PATH, encoding='utf-8'))
    print(f'questions:\n{questions}')
    return questions

def get_subcategories() -> dict:
    # Fetching all sub-categories from local: client\src\static-data\sub-categories.json
    subcategories = json.load(open(SUBCATEGORIES_JSON_PATH, encoding='utf-8'))
    print(f'subcategories:\n{subcategories}')
    return subcategories

def get_categories() -> dict:
    # Fetching all categories from local: client\src\static-data\categories.json
    categories = json.load(open(CATEGORIES_JSON_PATH, encoding='utf-8'))
    print(f'categories:\n{categories}')
    return categories

def get_answers(RNA_id) -> dict:

    # connect to an existing dynamodb
    dynamodb=boto3.resource('dynamodb',region_name=REGION)
    
    # connect to dynamodb tables
    rnas_table = dynamodb.Table(RNAS_TABLE_NAME) # decide between rnas, categories, subcategories
    answers_table = dynamodb.Table(ANSWERS_TABLE_NAME)  # depends on the Answers Table name 
    
    # GET RNA DATA
    rna_response = rnas_table.get_item(Key={'id':RNA_id})
    #rna_response = rnas_table.scan(FilterExpression=Attr(RNAS_PARTITION_KEY).eq(RNA_id))
    print(f"RNA:\n{rna_response['Item']}")
    
    # GET ANSWERS DATA
    ans_response = answers_table.scan(FilterExpression=Attr(ANSWERS_RNAS_TABLE_KEY).eq(RNA_id)) 
    answers = ans_response['Items']

    # Since scan can only retrieve 1MB of data at a time, we need to paginate to retrieve all data
    while 'LastEvaluatedKey' in ans_response:
        ans_response = answers_table.scan(FilterExpression=Attr(ANSWERS_RNAS_TABLE_KEY).eq(RNA_id),
                                      ExclusiveStartKey=ans_response['LastEvaluatedKey'])
        answers.extend(ans_response['Items'])
    
    print(f"answers:\n{answers}")

    # data = {}
    # answers = {}
    # for ans in response_data['Items']:
    #     answers[ans['questionId']] = {
    #                         'value': ans['value']['storedValue'],
    #                         'notes': ans['notes']
    #                         }
    # data[rna_id] = answers
    
    # print(f'[-] {rna_id} Done')

    # return data

def generate_excel_report(RNA_id):
    wb = Workbook()
    ws = wb.active
    
     # SET HEADER VALUES
    headers = ['Question #','Category', 'Subcategory', 'Question', 'Answer', 'Images Added']
    for i in range(len(headers)):
        ws.cell(row = 1, column = i+1).value = headers[i]
    
    categories = get_categories()
    subcategories = get_subcategories()
    questions = get_questions()
    answers = get_answers(RNA_id) 
    return

    row_num = 2
    for item in RNA_id:
        ws.cell(row= row_num, column = 1).value = item['rna']['id']
        ws.cell(row= row_num, column = 2).value = item['category']['name']
        ws.cell(row= row_num, column = 3).value = item['subcategory']['name']
        ws.cell(row= row_num, column = 4).value = item['question']['question']
        if item['question']['type'] == 'multi-select':
          answer_str = ''
          for answer in item['answer']['value']:
            if answer == item['answer']['value'][-1]:
               answer_str += answer
            else:
               answer_str += answer + ', '
          ws.cell(row= row_num, column = 5).value = answer_str
        else:
           ws.cell(row = row_num, column = 5).value = str(item['answer']['value'])
        ws.cell(row_num, 6).value = len(item['answer']['photos'])
        row_num += 1
    wb.save('report.xlsx')


def lambda_handler(event, context):

    try:   
        data = generate_excel_report(event['RNA_id'])

        return {
            'statusCode': 200,
            'body': json.dumps(data),
            'headers': {
				'Content-Type': 'application/json',
				'Access-Control-Allow-Origin': os.getenv("CORS")
			}
        }
    except ClientError as e:
        logger.error(f"Client error: {e}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error_message': 'Internal server error'}),
            'headers': {
				'Content-Type': 'application/json',
				'Access-Control-Allow-Origin': os.getenv("CORS")
			}
        }
    except Exception as e:
        logger.error(f"Error: {e}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error_message': 'Internal server error'}),
            'headers': {
				'Content-Type': 'application/json',
				'Access-Control-Allow-Origin': os.getenv("CORS")
			}
        }


# event = {'RNA_id': '831e5d34-26d0-4015-b1b7-7a3d33513830'}
# generate_excel_report(event['RNA_id'])