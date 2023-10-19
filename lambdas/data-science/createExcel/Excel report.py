from openpyxl import Workbook

JSON_example = [{
  "category": {
    "id": "5iezgjlw5l",
    "name": "General",
    "description": "4WS, Coordination, Demographics and more",
    "icon": "base64"
  },
  "subCategory": {
    "id": "edoevk4gpa",
    "name": "4WS",
    "categoryId": "5iezgjlw5l"
  },
  "question": {
    "title": "Are there other organizations in the area?",
    "hintText": "",
    "subCatergoryId": "edoevk4gpa",
    "type": "yes-no",
    "options": [],
    "dependencies": {
      "questionId": "c793tsbvm7",
      "expectedAnswer": "boolean"
    }
  },
  "rna": {
    "id": "6jvq9p4bzh",
    "status": "In Progress",
    "communityName": "Community Center",
    "communityType": "Combined",
    "lastSyncDate": "12/08/2023",
    "creationDate": "01/06/2023",
    "location": "Istanbul, Turkey",
    "version": "2023.06.01"
  },
  "answer": {
    "id": "b61ef06sll",
    "questionId": "edoevk4gpa",
    "rnaId": "ueiq8bb0r4",
    "value": "Yes",
    "photo": "base64",
    "otherText": "There are 5 more groups here"
  }
},
{
  "category": {
    "id": "7toxjz225o",
    "name": "Health",
    "description": "Services, Nutrition and more",
    "icon": "base64"
  },
  "subCategory": {
    "id": "36zg89cv9i",
    "name": "Mortality",
    "categoryId": "7toxjz225o"
  },
  "question": {
    "title": "Is there any specific problem regarding the mortality rates?",
    "hintText": "",
    "subCatergoryId": "36zg89cv9i",
    "type": "yes-no",
    "options": [],
    "dependencies": {
      "questionId": "hkafg2gr2e",
      "expectedAnswer": "boolean"
    }
  },
  "rna": {
    "id": "6jvq9p4bzh",
    "status": "In Progress",
    "communityName": "Community Center",
    "communityType": "Combined",
    "lastSyncDate": "12/08/2023",
    "creationDate": "01/06/2023",
    "location": "Istanbul, Turkey",
    "version": "2023.06.01"
  },
  "answer": {
    "id": "eessp1cj6v",
    "questionId": "hkafg2gr2e",
    "rnaId": "wpo1qjfglm",
    "value": "Yes",
    "photo": "base64",
    "otherText": "It happends only to children"
  }
},
{
  "category": {
    "id": "p6kpr4ql3d",
    "name": "WASH",
    "description": "Water, Sanitation, Hygiene and more",
    "icon": "base64"
  },
  "subCategory": {
    "id": "i45wygkdya",
    "name": "Hygiene",
    "categoryId": "p6kpr4ql3d"
  },
  "question": {
    "title": "What existing practices are harmful to health?",
    "hintText": "",
    "subCatergoryId": "i45wygkdya",
    "type": "multi-select",
    "options": ["Improper waste disposal", "Open defecation", "Unsafe cooking methods", "Limited hand hygiene",
    "Overconsumption of sugary foods", "Lack of vaccination awareness", "No significantly harmful practices observed"],
    "dependencies": {
      "questionId": "ab37783bvd",
      "expectedAnswer": "select-option"
    }
  },
  "rna": {
    "id": "6jvq9p4bzh",
    "status": "In Progress",
    "communityName": "Community Center",
    "communityType": "Combined",
    "lastSyncDate": "12/08/2023",
    "creationDate": "01/06/2023",
    "location": "Istanbul, Turkey",
    "version": "2023.06.01"
  },
  "answer": {
    "id": "n7ecxc1smk",
    "questionId": "ab37783bvd",
    "rnaId": "6jvq9p4bzh",
    "value": "[Open defecation, Unsafe cooking methods]",
    "photo": "base64",
    "otherText": ""
  }
}]

MIDUL_JSON = [{
  "category": {
    "id": "2",
    "name": "Education",
    "description": "",
    "iconSrc": "base64"
  },
  "subcategory": {
    "id": "3",
    "name": "General",
    "categoryid": "2"
  },
  "question": {
    "id": 11,
    "category": 2,
    "subcategory": 3,
    "order": 1,
    "question": "Are the schools shut down?",
    "answertype": "Yes or No",
    "options": [],
    "dependency": None
    },
  "rna": {
    "id": "6sah2013",
    "status": "50.3",
    "communityname": "shimshit",
    "communitytype": "communal village",
    "lastupdatedate": "12/08/2023",
    "creationdate": "12/08/2023",
    "location": "Shimshit, Israel"
  },
  "answer": {
    "id": "26812",
    "questionid": "11",
    "rnaid": "6sah2013",
    "value": ["Yes"],
    "photo": [],
    "otherText": ""
  }
},
{
  "category": {
    "id": 2,
    "name": "Education",
    "description": "",
    "iconSrc": "base64"
  },
  "subcategory": {
    "id": 4,
    "name": "Refugees/IDPs/Migrants",
    "categoryid": 2
  },
  "question": {
    "id": 35,
    "category": 2,
    "subcategory": 4,
    "order": 2,
    "question": "Is there a need for translation?",
    "answertype": "Yes or No",
    "options": [],
    "dependency": 34
    },
  "rna": {
    "id": "6sah2013",
    "status": "50.3",
    "communityname": "shimshit",
    "communitytype": "communal village",
    "lastupdatedate": "12/08/2023",
    "creationdate": "12/08/2023",
    "location": "Shimshit, Israel"
    },
  "answer": {
    "id": 209323,
    "questionid": 35,
    "rnaid": "6sah2013",
    "value": ["Yes"],
    "photo": [],
    "otherText": "string"
  }
},{
  "category": {
    "id": 1,
    "name": "Health",
    "description": "",
    "iconSrc": "base64"
  },
  "subcategory": {
    "id": 17,
    "name": "Immunizations",
    "categoryid": 1
  },
  "question": {
    "id": 239,
    "category": 1,
    "subcategory": 17,
    "order": 2,
    "question": "When did the last immunization occur?",
    "answertype": "Multiselection",
    "options": ["Within the last 6 months", "Within the last year", "Within the last 2 years", "Within the last 5 years"],
    "dependency": None
    },
  "rna": {
    "id": "6sah2013",
    "status": "50.3",
    "communityname": "shimshit",
    "communitytype": "communal village",
    "lastupdatedate": "12/08/2023",
    "creationdate": "12/08/2023",
    "location": "Shimshit, Israel"
  },
  "answer": {
    "id": 20492,
    "questionid": 239,
    "rnaid": "6sah2013",
    "value": ["Within the last year"],
    "photo": ["data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAbUAAABzCAMAAAAosmzyAAAA9lBMVEX////0mRoZYaz0lAAAWqkAVKfzkgAAWKj4w4ULXapkjcH97tz86dL0mAP5z536z6OduNjS3+6Fpc7J2OoAUqaPrNL+9+yYr9H85dD2rVr//fr2+fz0mADj7PX29vT+9Oc6d7f3tmj73bu+0OWwxd/v9Pn848j2rFL0nB93m8j4vXjg6fP5ypRUg7wlaK+qwNxBbZ82cbT1oDL2qkn74L9qk8X71674vXb2q01Pf7r1oir4xIz4vYD1oz33tmUAS6O/rZTAlF5QfKRQfK19m8PBy9QcXZ2drb/U2d+xvcvcnlJ/i402aqHKwbG7vLbclTd1iJzlsW9WqLUkAAARxUlEQVR4nO1daXviOBIGHziGOOFmA6EhARIIAYZ0SIf0JHN0T0/PHrO7///PLLYlWyWVDiDTpJ/1+2Ge6eBD1uuSql5Vyblchm+O3tXl6c3j/aGbkcEUzavb63E+sG278HDotmQwQunh7C6wg3wE+/TQzcmgR/PhJU8Zi1g7OXSLMmhxfM5StkFwdOgmZdDiGHK2sbUPh25SBi16HGn54OzQTcqgRZtnLf/+0E3KoIfA2lP70E3KoMWTMERmrL19PPOs2Rlrbx/XvBNZaB66SRm0uLB51kqHblIGLY541uyrQzcpgxYnAmuXh25SBi0eBNYy+fjt40pg7fbQTcqgxZUgRH48dJMyaCHKxz8eukkZtCiNedYuDt2kDFr0eEkrOMvEkbeC9oUkeG7ecazl32esvRXcFMZ4Gk9bECLvMknrjeChEE5YPeyn9zxr4y1ZGxZHr9HEDDxK55GfEWCyxzXP2vlWrA37njd5pWZmYNF+JKF04UVk5EUQ/VGTxLGYu75leYNXbGwGgpNC4tbnhdntw86i/6g896wIXjZGvjpASGZfcOa2q+ifcmZZTjZGvjpeAC/2E5zdRNHfJGe820o528CtGLWkM5oOhotyo1FeDAeVUWeHh/l/gag03rDmdimwps8+7jYAZ5blzyIGyjGG6EmdQXVt1b0NHCf8r+stiw1kRhy1ylJEZJs/+gCcvM0wfoJBthpyfMscdBt3bg+9ArnOw+VVSRMVCwFZ3r5jCmeOBdZ08vGo6jsWBzdiyongzpCTKhPHdXzuLN9x/QlPXOUnR4qQ7Pp8VZ5qmhijs/SYU+v4y4Ti2MYgS844KbAHlRRXYA4r3H14kHvrgi1FzHxIqBYkLftG+UDdiSdwtun/fmhs5P+LwkmVdR05KZ4TXW8CxteKKzmSIbverykbGWMABgS/b3AKgTDZK0ehU/Zoyho/wiGXswvPDxKTE1Ln4jPOE3O74K6vzBmvrFy8+yNjk7DWmQhWBonz1wwJetbCe3hFvb3N4F09s8k3xDnez3f40TuyFnX2+BYzuPuC7PgbQvMld0QgF/0rRV/W/dGbjLM27XuSkxgSUi/UiLUN1XNdlFjh2upUNSeknYaa2oaRY/TwPVgLeUOmSyGITg+/jlkWlEiZ6K/gLOyTgYS1ypw9y9/MMJtZhr+Sv0yPN2PN8v2Fuu+r/KjgmzqtQuIa7TJ8yt+LNSQYy5Uk94+OJvTwM997dJKsKTkL3UictWlKmu9Z89mk2io3qpPZcm45XvpTPzmBsuZj7ghog6u2NqG5nqE/0pR2Gp5OvydrefuZk6OEZB5wNPE7uFfrGWGtNnOVnIXoYqx1+vQ8x1sNu+wv01qj7xHiRNb8WUNAdbLquynVlq9y5odkXJ4vl0s/ebFMcCrttAJatb4va/lgzOlRx2OVtcWNaMM7PAlC5GCt58zyFhhrDTqnuVWsi0fDYuRdiqx5ZUmf1lYJb0pNZh0fVQ/flDlpg1HE0BYE9bTD0Dl/b9byeZ623I2CtiA+pBeAY7grDJcGnG26uY+w1k1+lo5m3epm5ERYa0m7dTqjr4Inp2FaZ+yrHJ9g5o+oZpU8Jq2/AmuBMMBd3UmbQZMfe++ZQ85Z1jrDuWfAme/Nyx2ENdJdlq+agkbVkHIKPWu53IpcVkED8UXiuWxER0sTfeSjanTC9JFXYA2x4iYeNIa4Qw5h5OPRYqn12kNO3P4Cj7LJrKaTl7srwYdUstahM1VfRsOIDIp+/M9ifLyncTtDNKFPbYO+C66RM/SsCdIIQpso/97LzC1dlrlKzC3JYx2Vl2Z21h/GPrXIGrUJ7YSSRsAmrOUGxKuXDpHD+CpOAxxv4o+AvN7g7OgIsBggIZuWNfvobwAnF895gZA70T9t8goIvV5q8e2TuyA6yI7fp5AzEztzKGcIa1NyheUWyq0Ra9S/kDrzRBdJXpclaZteHwEedVDiIiMsZNOzJkphxycBxxs6+D6gziTYqOLdz788Po3H+bvNlFvhZX2pnTEzlsBaLb5GrFIawoy1SWw8somtQm68pn+g/oh2JRDqsmGVOkxjQ1StXVgL0+Y4QsZYc5oviLgVvIBjvv76qTwclj+vLIncyHHmQhVXxtpfYGtEGfZX+M+E1DQ3okuMU+uPgAg3nmzAcn8gLhrvxprg9uDBYO7yXDS3R240/e3LT56jFnspHLfIjTcCaxWaomAu3BqyVnP4e7GgHC1TG18Z+iMgZ20cOfpgpkNCtl1Z44IyWa5+80ygTchZbf/+xWRotBxugSWEwFrHFf6khRlrFXJh3L1YkPGwkf6JDtZL9IQEJdDbMUNQqRW3G9iZNS6FeCxbKT3lZzck03j0WVz1FDizJoj3JlJE3nnLN9bbt2QNfx2ojMa0sUO0Eo3Z3wK7IskaH8VRk8XOrHGCozzPqsSZG5of/vWT2uP35hhnGGuJoOXMTNYxQxiOkCRgQ+c1+iOg1MwfAcUqT0SwKLEegRiy7c5aG9xOlbED03uuUats//FFbm6OVe1iJ2GsjZKz/Pp6ITkNYitvBOeA+CJwTWBEQzyVP3vF8pOs6rfBZCckIuzOWu5H4Oiodp4rMSF38HfJM7z7HUs0CB/aaUjdMGTUWqSTpO/V/VmrpgsCtvL8UY2ZzKZ8vKE6hQL4B4XEXQS8CBzswRqIBRXr0SFuk+0Fg1+kFLyb1MVh0rPKil5HWMtNYPLG5l3or8o1hdWZsUaiZnSWKpNYjrtCjejJfflV2yD1IF1Na4Kg6Zkbn/ZgDYy9unKmYypfnf/jnfyoCreo5nvLhdJSMNY6K35t2t9QN+9PFhWcuq0UrTl2DSJSevxvhGlXPsPCYI2RQcBmOjanau3BWg6w9qQpsWgTX+nxBwVrm86ZeTR28x2vr+ZMkoHQaVhYUpfjefNiuSJe0Yi1pUKVrpEBUnBU9P7IBdL/ESCdXGC1D2vAuPXlTFfPm3ud//xPJWu5TqXan2862J+vq9oZSZajNS1ivEXUuWshuXSblRo05xn1RUKMfIWBhuiBPmS1qyaQcbkSv31YA/pZoC9nar7YwfUnDWshppVaTTKacZCwtrlE1Zc4N77nrmDXb7EqisbYlBtL/Imu18j8kRO50wHS7m2oPe3DGgjgzw3Kmdov5/+qq0fILSFlbYNBde7iUaBfn7D2ps9ASGbbOvYukXEQY53KYH38yjD1IABdCJLtOPX29VgzKR1s3y/9b8baBtPFZFnfjLYCd47FmJsi26cxWSX5QSFpqNtPdBEfY5Qu0uL+CEyefwS/tYH4BNM8Xm2EzBsVfK6cb8paLqyqGZQnfUewOi/tR+PMOnwMpXoj2oiy6kcYrPGFtSAfAK6FfUNvZIOWZ31r1iKMuoNGcelB5hKJzDCL1bdw9Z7MXT5qTjQvwcIMkdNyH88AYP36M3viHqy1YSBowFpjM8obsNbsHd/fl0omdb+GrEUYVRprixFOZtRDNczzX+Mq8JSuow1qGGYKf4TP5g0AuN/YkG0P1o6Bwqnf6mVUDLtMx1r76sfnoGDbhWB8Jqv/SLENayGmTJJD4qgbsOY4fVnOVytJunMRJOY9R04Vl7HkAIkIe7B2oooDRQzmkT+lYe3ysZBcNrCfTjS8bctajtEpk7P0rC0b8tWWue5kwqnIuvBpAyUC5sw9WAPjrq7eszOpk75SsVZ65HIW7PNbJW87sJYkBCWxL2XNESyFRHyOIl1/YJI4YaErPPI8cQxsyLY7a1CGlJTsJM82pw+nYO3dUUHM67LtI8UEtwtruQVpC9U5CGtOdVqBmCaLdXIhcWaUOxHeTVDN8UI/KWtMld/urMGKJ+ladojOJH0hpay9+/Pfkly8/AfpkutOrHVIU2iSnEIbKZIoWZo81DWsohIXBLjX3oC2tIt3Zg0uZSuLdGtLdhRZo+Ji+4//SEhT8rYTazT0pX6dgrWurjiG+iIeFusRkAdfcg+uyhNHeyFdeN6VtXt4xYJ8x+lOAyRg4ZmKXz/98KzKU7eDD+gQvBtrRWPWqCRleQ3kx1zii8zluyiUy30yREJ/pM3vk6kD8xmmHVk75aMJaQ9VuGpbjLV3nx3/v5riAtu+QHjbz9b0I+TGmMgQWEc9/wGdElW3G6KZlNJKaHkXJBP8TqwdX/O5x1IPssErSSJr7d8348ufknJy9iaFFyGfk2etuzYoGEvmNeiNSDR/MrVZPnZlYrV15V07PnYFuDFVAQechpKQTc/abROgdH9yJ7wlMl9kKibu+/y89tsXz/I/vZgM8kHhPZctK7D2k7fS5q8uuGVpNWsdWiuDDBJTmp+uviGtkWJH2SZ8MsmZUD1J8gVMKqGg0IIU1ci+o9BCFkq4ef3rr5Fg8ts1dl0E9h24l8Ba3XL8ojqljmZx8VG2bH2NJqEjK9KgZE0OujcC64/AtWqpMwcfns4Rr1K/hu8XUlljKcXg2UefLd+KSxraVzeq0mDmbs/MDjUIa6F2vx7Kk+u7JLk0VSt0q6ItGmzzWuKI5IVYujV3ms/KzI1nbCcjqfwEYFUg6ejXqBW9Q8PgsoXGn8yjdxZzMHT0jkx5O6W8oayFRRzLyQDtys096SzT59Rj+Vp2MrVxRjw0LZuh/kg6zoA88fyT9MwrNBHhNWwN88m7M0nufpp4MVjTCDXJVut9fDIcJ4lAKWEt7ErXKrYGnMlVWulEmzqFWtZGS/oCwusRXURfopYYZeKPwDxxxUZiYOmZTkX7s4bu7jiUbhVCte9RuqzPJsM0TwztbXwSvndy1qwoK6vur1fV8mI4HC7Kk7XPVFt5qbOuzxuhU1tanRad5yN/xMH7I21tOSgFoJfmju+930iAqBajmVTnSSJbJokK1jq3T/maRsmdzz82ER+Sv6HvR7sMeo4Hgn2PCfEMcrTKrsC1IjVLBFWsSdE2N/A9Ks6EQynZLWLfvX2ukcXQhWJPHrJH0YClNV2fJHiQ76HAonArRtkLo8Juy2NvaZJZR6c2xl2kYZ9RgWMR+iMgDVK9q+kZEj2//j5a3aKiGo08dAUmdQvSRvvy0YS3HqKNdBYzbVmV71fZ98SEtSSTIE0ToiVrRiVXQ5pRFP0LBmtj5ao9FHxjv2WfPevy2BrKaK7oMy8WddLZPX4UpLiofX9t6zSuF9nuZxPljiW+t4a+oFHuMRUk0+Gcpl8Zbd/TAZv9wGBNXSLRBHplvCK2K2t28HyLvyKKBSeHjEtVaI2457zhTT3BhUEO2fWUN9bRcMbudAYo85eSvVhddXVGwyUbvxbBWcK9JajGp7uRZb4Hm6RqvvV4AQ6OapdOd9iLNQjG7z8eyxbUBnJPZBa/phXOGh2Jnp67UvIW6jvVGEjyVKfWKlp11934IX6EcDdc1ykuRDe9O4mvovEqWuRmpM54QP5pWNvYJYc3whf35oiFJsviGBwcsXbF/uEmtp7ekRy3Jw+X91LGIixxzqw6HQgnPGvy+q7j64KMN90rGqIzHSzK1VVxti6uJtXW0Gz34v9LDFFjcxz6Knf5/BiljFf6gAuUAV5zmmFXLJEJxUu/CDTgfUzN90t6PyL7XzBllRleBWLakjNnZp4G/7P2iwqIQGm/aM7JsC14NxJuz13kTdHVFz31jjiBEpNkMuwFuNe270NvY82z5piIC71b1t6yrzX/BWgwU5fX5zy3vjDtme151WQ2XUP37cqwHzqJPflC7NoRIgPjncradKMgw69IZdgOVOP2+oKnMeJZ22pXwIfngsGHbTLshkUYtPlYFqEQrgmSvxLty8cCv+1dhtfCytkYGqb2VPgBcttsxvZl5j/+VegsvQk6X/Efc8m+W/iWUJEoHjU+yDb/Bk+Gg0EQtKSSf4a3g4UgQ6q2dMvwNtASWNviQ44ZDgThI2XZJ8u/A6wE8XiL7cAzHAhCXomB5J/h0BDE43r27fG3D2F/Didj7e2DJ82aZ6y9eYwE1vqHblIGLbr8tGZSkpLhwBDE420l/wwHgCgeZ5L/24eQdpeJx98Bhpl4/B1iyO9AbPKV2wyHxrS19lniMsn/+0Cn0lrPkzJpk7LmDG8D3UF1PbfCEvd6Jvl/X5jWhq1Jfw/J/389XXzqj653EwAAAABJRU5ErkJggg=="],
    "otherText": "string"
  }
},{
  "category": {
    "id": 5,
    "name": "WASH",
    "description": "",
    "iconSrc": "base64"
  },
  "subcategory": {
    "id": 9,
    "name": "Sanitation Excreta disposal",
    "categoryid": 5
  },
  "question": {
    "id": 109,
    "category": 5,
    "subcategory": 9,
    "order": 1,
    "question": "Are people unfamiliar with the design, construction and use of toilets?",
    "answertype": "Yes Or No",
    "options": [],
    "dependency": None
    },
    "rna": {
      "id": "6sah2013",
      "status": "50.3",
      "communityname": "shimshit",
      "communitytype": "communal village",
      "lastupdatedate": "12/08/2023",
      "creationdate": "12/08/2023",
      "location": "Shimshit, Israel"
    },
  "answer": {
    "id": 29182,
    "questionid": 109,
    "rnaid": "6sah2013",
    "value": ["Yes"],
    "photo": [],
    "otherText": ""
  }
},{
  "category": {
    "id": 5,
    "name": "WASH",
    "description": "",
    "iconSrc": "base64"
  },
  "subcategory": {
    "id": 9,
    "name": "Sanitation Excreta disposal",
    "categoryid": 5
  },
  "question": {
    "id": 119,
    "category": 5,
    "subcategory": 9,
    "order": 3,
    "question": "Are the latrines or toilets  not cleaned and maintained resulting  they are not hygienic and safe for all users",
    "answertype": "Yes Or No",
    "options": [],
    "dependency": 109
    },
    "rna": {
      "id": "6sah2013",
      "status": "50.3",
      "communityname": "shimshit",
      "communitytype": "communal village",
      "lastupdatedate": "12/08/2023",
      "creationdate": "12/08/2023",
      "location": "Shimshit, Israel"
    },
  "answer": {
    "id": 28721,
    "questionid": 119,
    "rnaid": "6sah2013",
    "value": ["Yes"],
    "photo": [],
    "otherText": ""
  }
}]

def generate_excel_report(json):
    wb = Workbook()
    ws = wb.active
    headers = ['RNA_id', 'category', 'subcategory', 'question', 'answer', 'images_number']
    for i in range(len(headers)):
        ws.cell(row = 1, column = i+1).value = headers[i]
    row_num = 2
    for item in json:
        ws.cell(row= row_num, column = 1).value = item['rna']['id']
        ws.cell(row= row_num, column = 2).value = item['category']['name']
        ws.cell(row= row_num, column = 3).value = item['subcategory']['name']
        ws.cell(row= row_num, column = 4).value = item['question']['question']
        ws.cell(row= row_num, column = 5).value = item['answer']['value'][0]
        ws.cell(row_num, 6).value = len(item['answer']['photo'])
        row_num += 1
    wb.save('report.xlsx')


generate_excel_report(JSON_example)



