import os
import uvicorn
import pytz
import json
import pandas as pd
from transformers import AutoTokenizer
from transformers import AutoModelForSequenceClassification
from transformers import pipeline
from scipy.special import softmax
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle, numbers
from datetime import datetime
from fastapi import FastAPI, Request, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from sentence_transformers import SentenceTransformer, util

def current_time() -> str:
    local_tz = pytz.timezone('Asia/Jerusalem')
    date = datetime.now().astimezone(tz=local_tz).strftime("%d/%m/%Y %H:%M:%S")
    return datetime.strptime(date, "%d/%m/%Y %H:%M:%S")

def run_uvicorn() -> None:
    print(f'{current_time()} | Starting uvicorn server  ...')
    uvicorn.run(app=f'{os.path.basename(__file__).split(".py")[0]}:app', host='0.0.0.0', port=8080, log_level='info', reload=True)

def get_excel(payload):
    """
    Create excel report with all of the questions and answers, without pictures.
    """
    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~ EXCEL VARIABLES ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    header = NamedStyle(name="header")
    header.font = Font(name='Calibri', size=12, bold=True, color='000000')
    header.border = Border(top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        left=Side(border_style="thin")
                        )
    header.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    header.alignment = Alignment(horizontal="center", vertical="center")
    header.fill = PatternFill(start_color='CADBFF', fill_type='solid')

    title = NamedStyle(name="title")
    title.font = Font(name='Calibri', size=16, bold=True, color='000000')
    title.border = Border(right=Side(border_style="thin"),
                        left=Side(border_style="thin")
                        )
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.fill = PatternFill(start_color='E7EBE6', fill_type='solid')

    entry = NamedStyle(name="entry")
    entry.font = Font(name='Calibri', size=12, bold=False, color='000000')
    entry.border = Border(bottom=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        left=Side(border_style="thin")
                        )
    entry.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    entry.alignment = Alignment(horizontal="center", vertical="center")


    wb_name = f'{payload[0]["rna"]["id"]} report.xlsx'

    # CREATE INITIAL EXCEL FILE WITH ALL OF THE QUESTIONS AND ANSWERS
    df = pd.json_normalize(payload)
    df_wanted = df[['category.name','subCategory.name','question.title','answer.value','answer.otherText','answer.photo']]
    df_wanted.to_excel(wb_name, index=False, header=False)

    # SET WORKSHEET
    wb = load_workbook(wb_name)
    ws = wb.active
    
    # ADD LOGO
    # israaid_logo = drawing.image.Image("./elements/israaid.png")
    # israaid_logo.height = 60

    # SET HEADER VALUES
    ws['A1'] = 'Category'
    ws['B1'] = 'Subcategory'
    ws['C1'] = 'Question'
    ws['D1'] = 'Answer'
    ws['E1'] = 'More Details'
    ws['F1'] = 'Image_Attached'

    # TITLE STYLE
    ws.insert_rows(1)
    ws.merge_cells('A1:F1')
    cell = ws.cell(row=1, column=1)
    ws.row_dimensions[1].height = 60
    cell.value = f'Created: {df.iloc[0]["rna.creationDate"]} | Updated: {df.iloc[0]["rna.lastSyncDate"]} | Location: {df.iloc[0]["rna.communityName"]}, {df.iloc[0]["rna.location"]}'
    cell.style = title
    # israaid_logo.anchor = 'H1'
    # ws.add_image(israaid_logo)

    # HEADERS STYLE
    header_row = ws[2]
    for cell in header_row:
        cell.style = header
    ws.row_dimensions[2].height = 20

    # ADD BORDERS TO ALL CELLS
    for row in range(3, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=column)
            cell.style = entry

    wb.save(wb_name)
    wb.close()


app = FastAPI()
origins = ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

### Define the API endpoints ###  
@app.get("/")
async def root():
    raise HTTPException(status_code=200, detail="IsraAID API is Running")
    
@app.post("/make_excel")
async def make_excel(request: Request):
    """
    Create excel report with all of the questions and answers, without pictures.
    """
    
    payload = await request.json()
    payload = str(payload).replace("'", '"') # convert single quotes to double quotes because of python json format

    payload = json.loads(payload)
    # print(type(payload),payload)
    # print(type(payload))
    get_excel(payload)

    return HTTPException(status_code=200, detail="Excel created successfully", headers=None)

@app.post("/similarity")
async def similarity(request: Request):
    """ 
    Compute similarity between each sentence in docs list and others list
    Using model: https://huggingface.co/sentence-transformers/all-MiniLM-L6-v2

    Parameters
    ----------
    Accepts list of new question and historic questions [[new],[historic]] 
    
    Returns
    -------
    list of similarity scores between docs and others
    """

    payload = await request.json()

    transformer_model = SentenceTransformer('sentence-transformers/all-MiniLM-L6-v2')

    #Compute embedding for both lists
    origin_embedding = transformer_model.encode(payload['docs'], convert_to_tensor=True, show_progress_bar=True)
    other_embedding = transformer_model.encode(payload['others'], convert_to_tensor=True, show_progress_bar=True)
    
    sim = util.pytorch_cos_sim(origin_embedding, other_embedding).tolist()

    return HTTPException(status_code=200, detail=sim, headers=None)

@app.post("/severity")
async def severity(request: Request):
    """ 
    sentiment analysis using roberta model from huggingface
    Using model: https://huggingface.co/cardiffnlp/twitter-roberta-base-sentiment-latest

    Parameters
    ----------
    Accepts text to be analyzed
    
    Returns
    -------
    only negative sentiment score (How bad is it)
    """

    payload = await request.json()

    roberta = "cardiffnlp/twitter-roberta-base-sentiment-latest"
    tokenizer = AutoTokenizer.from_pretrained(roberta)    
    model = AutoModelForSequenceClassification.from_pretrained(roberta)

    # sentiment analysis with roberta
    encoded_tweet = tokenizer(payload["answer"]["otherText"], return_tensors='pt')
    output = model(**encoded_tweet)

    scores = output[0][0].detach().numpy()
    neg_score, neu_score, pos_score = softmax(scores)

    return HTTPException(status_code=200, detail=str(neg_score), headers=None)

@app.post("/ask")
async def ask(request: Request):
    """ 
    This is a Question and Answering model about any context given
    Using model: # https://huggingface.co/distilbert-base-uncased-distilled-squad?text=Which+name+is+also+used+to+describe+the+Amazon+rainforest+in+English%3F&context=The+Amazon+rainforest+%28Portuguese%3A+Floresta+Amaz%C3%B4nica+or+Amaz%C3%B4nia%3B+Spanish%3A+Selva+Amaz%C3%B3nica%2C+Amazon%C3%ADa+or+usually+Amazonia%3B+French%3A+For%C3%AAt+amazonienne%3B+Dutch%3A+Amazoneregenwoud%29%2C+also+known+in+English+as+Amazonia+or+the+Amazon+Jungle%2C+is+a+moist+broadleaf+forest+that+covers+most+of+the+Amazon+basin+of+South+America.+This+basin+encompasses+7%2C000%2C000+square+kilometres+%282%2C700%2C000+sq+mi%29%2C+of+which+5%2C500%2C000+square+kilometres+%282%2C100%2C000+sq+mi%29+are+covered+by+the+rainforest.+This+region+includes+territory+belonging+to+nine+nations.+The+majority+of+the+forest+is+contained+within+Brazil%2C+with+60%25+of+the+rainforest%2C+followed+by+Peru+with+13%25%2C+Colombia+with+10%25%2C+and+with+minor+amounts+in+Venezuela%2C+Ecuador%2C+Bolivia%2C+Guyana%2C+Suriname+and+French+Guiana.+States+or+departments+in+four+nations+contain+%22Amazonas%22+in+their+names.+The+Amazon+represents+over+half+of+the+planet%27s+remaining+rainforests%2C+and+comprises+the+largest+and+most+biodiverse+tract+of+tropical+rainforest+in+the+world%2C+with+an+estimated+390+billion+individual+trees+divided+into+16%2C000+species

    Parameters
    ----------
    Accepts a request={"question","context", "threshold"}
    
    Returns
    -------
    string of the answer
    """

    payload = await request.json()

    question_answerer = pipeline("question-answering", model='distilbert-base-uncased-distilled-squad')

    response = question_answerer(question=payload["question"], context=payload["context"])
    
    answer = f"{response['answer']} | Reliability: {round(response['score'], 4)}"
    return HTTPException(status_code=200, detail=answer, headers=None)

if __name__ == '__main__':
    run_uvicorn()
