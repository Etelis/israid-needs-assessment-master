import os
import glob
import json
from json import load
import logging
import boto3 
import datetime
import urllib
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle, numbers
from botocore.exceptions import ClientError
from boto3.dynamodb.conditions import Key, Attr

## Change this global settings if necessary
REGION = 'eu-north-1'
ANSWERS_TABLE_NAME = 'Answers'
ANSWERS_RNAS_TABLE_KEY = 'rnaId'
RNAS_TABLE_NAME = 'Rnas'
RNAS_PARTITION_KEY = 'id'
CATEGORIES_JSON_PATH = "categories.json"
SUBCATEGORIES_JSON_PATH = "sub-categories.json"
QUESTIONS_JSON_PATH = "questions.json"

ISRAAID_BUCKET = 'israaidbucket'

## IsraAID logo link
LOGO_URL = 'https://upload.wikimedia.org/wikipedia/commons/thumb/c/c4/Logo-Israaid.svg/2560px-Logo-Israaid.svg.png'

### ~~~~~~~~~~~~~~~~~~~~~~~~~~~ EXCEL VARIABLES ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
header = NamedStyle(name="header")
header.font = Font(name='Calibri', size=11, bold=False, color='000000')
header.border = Border(top=Side(border_style="thin"),
                       bottom=Side(border_style="thin"),
                       right=Side(border_style="thin"),
                       left=Side(border_style="thin")
                       )
header.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
header.alignment = Alignment(horizontal="center", vertical="center")
header.fill = PatternFill(start_color='CADBFF', fill_type='solid')

title = NamedStyle(name="title")
title.font = Font(name='Segoe UI Bold', size=20, bold=True, color='000000')
title.border = Border(right=Side(border_style="thin"),
                      left=Side(border_style="thin")
                      )
title.alignment = Alignment(horizontal="center", vertical="center")
title.fill = PatternFill(start_color='E7EBE6', fill_type='solid')

subtitle = NamedStyle(name="subtitle")
subtitle.font = Font(name='Segoe UI Light', size=14, bold=False, color='000000')
subtitle.border = Border(right=Side(border_style="thin"),
                      left=Side(border_style="thin")
                      )
subtitle.alignment = Alignment(horizontal="center", vertical="center")
subtitle.fill = PatternFill(start_color='E7EBE6', fill_type='solid')

entry = NamedStyle(name="entry")
entry.font = Font(name='Calibri', size=11, bold=False, color='000000')
entry.border = Border(bottom=Side(border_style="thin"),
                      right=Side(border_style="thin"),
                      left=Side(border_style="thin")
                      )
# entry.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
entry.alignment = Alignment(horizontal="center", vertical="center")

### ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

answers_example = [
            {
            'photos': [], 
            'rnaId': '831e5d34-26d0-4015-b1b7-7a3d33513830', 
            'value': 
                    {
                    'storedValue': ['health services', 'education institutions', 'shelters']
                    }, 
            '_md': '2023-11-17T15:28:52.955Z', 
            'questionId': '40', 
            'createdOn': '2023-11-17T15:27:26.265Z', 
            'notes': None, 
            'id': '2c2714f8-470c-44c5-b94a-e753227fa601', 
            '_et': 'Answer', 
            '_ct': '2023-11-17T15:28:52.955Z'
            },
            {
            'photos': ['831e5d34-26d0-4015-b1b7-7a3d33513830','831e5d34-26d0-4015-b1b7-7a3d33513830','831e5d34-26d0-4015-b1b7-7a3d33513830'], 
            'rnaId': '831e5d34-26d0-4015-b1b7-7a3d33513830', 
            'value': 
                    {
                    'storedValue': ['health services', 'education institutions', 'shelters']
                    }, 
            '_md': '2023-11-17T15:28:52.955Z', 
            'questionId': '3', 
            'createdOn': '2023-11-17T15:27:26.265Z', 
            'notes': 'There are insufficient funds and lack of resources to provide the necessary services', 
            'id': '2c2714f8-470c-44c5-b94a-e753227fa601', 
            '_et': 'Answer', 
            '_ct': '2023-11-17T15:28:52.955Z'
            }, 
            {
            'photos': [], 
            'rnaId': '831e5d34-26d0-4015-b1b7-7a3d33513830', 
            'value': 
                    {
                    'storedValue': True
                    }, 
            '_md': '2023-11-17T14:42:37.642Z', 
            'questionId': '39', 
            'createdOn': '2023-11-17T14:42:25.947Z', 
            'notes': 'Note that This question is not relevant to the current situation', 
            'id': '4b225160-bdb0-4faa-aae1-b487351d6f30', 
            '_et': 'Answer', 
            '_ct': '2023-11-17T14:42:37.642Z'
            }
            ]

rna_example = {
        'emergencies': ['Earthquake'], 
        'location': '', 
        'affectedHouseholds': 300, 
        'creatorName': 'Or Avnon', 
        '_et': 'RNA', 
        'communityType': 'Hospital', 
        '_ct': '2023-11-01T00:11:39.454Z', 
        'isCompleted': False, 
        'creatorMail': 'turtcbui123@gmail.com', 
        '_md': '2023-11-01T00:11:39.454Z', 
        'createdOn': '2023-10-30T17:38:17.223Z', 
        'communityName': 'Peru Central', 
        'id': '831e5d34-26d0-4015-b1b7-7a3d33513830'
        }


# Configure logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Search for the category of the answer
def search_category(categories: dict, category_id: str) -> str:
    '''
    Search for the category of the answer
    ------- 
    Parameters:
        categories (dict): All categories
        category_id (str): The id of the category to search
    ------- 
    Returns:
        str: The category of the answer
    '''
    for cat in categories:
        if cat['id'] == category_id:
            return cat['name']

# Find the subcategory and category of the answer
def search_subcategory(subcategories: dict, subcategory_id: str) -> tuple:
    '''
    Find the subcategory and category of the answer
    ------- 
    Parameters:
        subcategories (dict): All subcategories
        subcategory_id (str): The id of the subcategory to search
    ------- 
    Returns:
        tuple: The subcategory and category of the answer
    '''
    for sub in subcategories:
        if sub['id'] == subcategory_id:
            subcategory = sub['name']
            category_id = sub['categoryId']
            return (subcategory, category_id)

# Find the question and subcategory of the answer
def search_question(questions: dict, question_id: str) -> tuple:
    '''
    Find the question and subcategory of the answer
    -------
    Parameters:
        questions (dict): All questions
        question_id (str): The id of the question to search
    ------- 
    Returns:
        tuple: The question and subcategory of the answer
    '''
    for q in questions:
        if q['id'] == question_id:
            question = q['title']
            subcategory_id = q['subCategoryId']
            return (question, subcategory_id)

# Formulate an answer with bool and text
def formulate_answer(answer: dict) -> str:
    '''
    Formulate an answer with bool and text
    -------
    Parameters:
        answer (dict): The answer to formulate
    -------
    Returns:
        str: The formulated answer
    '''
    added_notes = ''
    if type(answer['value']['storedValue']) == bool:

        if answer['value']['storedValue'] == True:
            answer['value']['storedValue'] = "Yes"
        else:
            answer['value']['storedValue'] = "No"
        
        if answer['notes'] != None:
            added_notes = f", {answer['notes']}"  

    elif type(answer['value']['storedValue']) == list:
        
        if answer['notes'] != None:
            added_notes = f", {answer['notes']}"    
        
        
    return f"{answer['value']['storedValue']}{added_notes}"

# Download the IsraAID logo and save it locally
def download_image() -> str:
    '''
    Download the IsraAID logo and save it locally
    -------
    Returns:
        str: The local path of the downloaded image
    '''
    try:
        local_path = os.path.join("/tmp/","logo.png")
        urllib.request.urlretrieve(LOGO_URL, local_path)
    except Exception as e:
        print(f"An error occurred: {e}")
    
    return local_path

def get_all_questions() -> dict:
    # Fetching all questions from local: client\src\static-data\questions.json
    questions = json.load(open(QUESTIONS_JSON_PATH, encoding='utf-8'))
    return questions

def get_all_subcategories() -> dict:
    # Fetching all sub-categories from local: client\src\static-data\sub-categories.json
    subcategories = json.load(open(SUBCATEGORIES_JSON_PATH, encoding='utf-8'))
    return subcategories

def get_all_categories() -> dict:
    # Fetching all categories from local: client\src\static-data\categories.json
    categories = json.load(open(CATEGORIES_JSON_PATH, encoding='utf-8'))
    return categories

def get_rna(RNA_id) -> dict:

    # connect to an existing dynamodb
    dynamodb=boto3.resource('dynamodb',region_name=REGION)
    
    # connect to dynamodb tables
    rnas_table = dynamodb.Table(RNAS_TABLE_NAME) # decide between rnas, categories, subcategories
    
    # GET RNA DATA
    rna_response = rnas_table.get_item(Key={'id':RNA_id})['Item']
    #rna_response = rnas_table.scan(FilterExpression=Attr(RNAS_PARTITION_KEY).eq(RNA_id))

    return rna_response
    
def get_rna_answers(RNA_id) -> dict:

    # connect to an existing dynamodb
    dynamodb=boto3.resource('dynamodb',region_name=REGION)
    
    # connect to dynamodb tables
    answers_table = dynamodb.Table(ANSWERS_TABLE_NAME)  # depends on the Answers Table name 
        
    # GET ANSWERS DATA
    ans_response = answers_table.scan(FilterExpression=Attr(ANSWERS_RNAS_TABLE_KEY).eq(RNA_id)) 
    answers = ans_response['Items']

    # Since scan can only retrieve 1MB of data at a time, we need to paginate to retrieve all data
    while 'LastEvaluatedKey' in ans_response:
        ans_response = answers_table.scan(FilterExpression=Attr(ANSWERS_RNAS_TABLE_KEY).eq(RNA_id),
                                      ExclusiveStartKey=ans_response['LastEvaluatedKey'])
        answers.extend(ans_response['Items'])
      
    return answers

def generate_excel_report(RNA_id):

    answers = get_rna_answers(RNA_id) 
    rna = get_rna(RNA_id) 
    categories = get_all_categories()
    subcategories = get_all_subcategories()
    questions = get_all_questions()

    wb = Workbook()
    ws = wb.active

    # TITLE INSERT AND STYLE
    # ws.insert_rows(1)
    ws.merge_cells('A1:F1')
    cell = ws.cell(row=1, column=1)
    ws.row_dimensions[1].height = 45
    cell.value = f'{rna["communityName"]} | {rna["communityType"]} | {rna["location"]}'
    cell.style = title

    # SUBTITLE INSERT AND STYLE
    # ws.insert_rows(1)
    ws.merge_cells('A2:F2')
    cell = ws.cell(row=2, column=1)
    ws.row_dimensions[2].height = 35
    cell.value = f'Created by: {rna["creatorName"]}, Last Update: {rna["createdOn"]}'
    cell.style = subtitle

    # LOGO INSERT - if it doesnt work, skip
    try:
        logo_img = openpyxl.drawing.image.Image(download_image())
        logo_img.width = 500
        logo_img.height = 70
        logo_img.anchor = 'H1'
        ws.add_image(logo_img)
    except:
        pass

    # SET HEADER VALUES
    ws['A3'] = 'Question #'
    ws['B3'] = 'Category'
    ws['C3'] = 'Subcategory'
    ws['D3'] = 'Question'
    ws['E3'] = 'Answer'
    ws['F3'] = 'Images Added'

    # HEADERS STYLE
    header_row = ws[3]
    for cell in header_row:
        cell.style = header
    ws.row_dimensions[3].height = 20
    
    # POPULATE DATA
    row_num = 4
    for ans in answers:
        answer = formulate_answer(ans)
        (question, subcategory_id) = search_question(questions, ans['questionId'])
        (subcategory, category_id) = search_subcategory(subcategories, subcategory_id)
        category = search_category(categories, category_id)
        
        # insert question number
        ws.cell(row= row_num, column = 1).value = ans['questionId']
        # insert category
        ws.cell(row= row_num, column = 2).value = category
        # insert subcategory
        ws.cell(row= row_num, column = 3).value = subcategory
        # insert question
        ws.cell(row= row_num, column = 4).value = question
        # insert answer
        ws.cell(row= row_num, column = 5).value = answer
        # insert images added
        ws.cell(row= row_num, column = 6).value = int(len(ans['photos']))

        row_num += 1
        
    # ADD BORDERS TO ALL CELLS
    for row in range(4, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=column)
            cell.style = entry
    
    wb_name = f'report_{rna["communityName"]}_{datetime.datetime.now().strftime("%d_%m_%Y")}.xlsx'
    wb_path = os.path.join("/tmp/", wb_name)
    wb.save(wb_path)

    return wb_name, wb_path

def lambda_handler(event, context):

    try:
        (wb_name, wb_path) = generate_excel_report(event['RNA_id'])
        
        # Return the file as response
        s3 = boto3.client('s3')
        bucket_name = ISRAAID_BUCKET
        s3.upload_file(wb_path, bucket_name, wb_name)

        # Get the URL of the uploaded file
        file_url = s3.generate_presigned_url(
                    'get_object',
                    Params={'Bucket': bucket_name, 'Key': wb_name},
                    ExpiresIn=3600  # Link expiration time in seconds
                    )

        return {
            'statusCode': 200,
            'body': file_url,
            'headers': {
				'Content-type' : "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
				'Access-Control-Allow-Origin': os.getenv("CORS")
			            }
        }
    except ClientError as e:
        logger.error(f"Client error: {e}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error_message': 'Internal server error'}),
            'headers': {
				'Content-Type': 'application/json',
				'Access-Control-Allow-Origin': os.getenv("CORS")
			}
        }
    except Exception as e:
        logger.error(f"Error: {e}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error_message': 'Internal server error'}),
            'headers': {
				'Content-Type': 'application/json',
				'Access-Control-Allow-Origin': os.getenv("CORS")
			}
        }


# event = {'RNA_id': '831e5d34-26d0-4015-b1b7-7a3d33513830'}
# data = generate_excel_report(event['RNA_id'])
